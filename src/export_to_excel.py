"""
豆瓣电影数据导出到Excel模块
作者: mshellc
"""

import json
import pandas as pd
import os
from datetime import datetime
import requests
import hashlib
import argparse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

def download_image(url, timeout=5, cache_dir='image_cache'):
    """下载图片并返回BytesIO对象，同时缓存到本地文件夹"""
    if not url:
        return None
    
    # 创建缓存目录
    os.makedirs(cache_dir, exist_ok=True)
    
    # 生成缓存文件名（使用URL的MD5哈希）
    import hashlib
    url_hash = hashlib.md5(url.encode()).hexdigest()
    cache_path = os.path.join(cache_dir, f"{url_hash}.jpg")
    
    # 检查缓存是否存在
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'rb') as f:
                print(f"使用缓存图片: {url}")
                return io.BytesIO(f.read())
        except Exception as e:
            print(f"读取缓存失败 {url}: {e}")
    
    try:
        print(f"开始下载图片: {url}")
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://movie.douban.com/'
        }
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        
        # 检查是否为图片
        content_type = response.headers.get('content-type', '')
        if not content_type.startswith('image/'):
            print(f"警告: {url} 不是图片文件 (Content-Type: {content_type})")
            return None
        
        img_data = response.content
        
        # 保存到缓存
        try:
            with open(cache_path, 'wb') as f:
                f.write(img_data)
            print(f"图片已缓存: {cache_path}")
        except Exception as e:
            print(f"保存缓存失败 {url}: {e}")
            
        print(f"图片下载完成: {url}")
        return io.BytesIO(img_data)
    except requests.exceptions.Timeout:
        print(f"警告: 下载图片超时 {url}")
        return None
    except requests.exceptions.RequestException as e:
        print(f"警告: 下载图片失败 {url}: {e}")
        return None
    except Exception as e:
        print(f"警告: 下载图片时发生未知错误 {url}: {e}")
        return None

def export_douban_to_excel(use_latest_only=True, include_images=True):
    """从data目录导出豆瓣电影数据到Excel
    
    Args:
        use_latest_only: 是否只使用最新的JSON文件，避免处理过多数据
        include_images: 是否包含封面图片
    """
    
    # 读取data目录下的JSON文件
    data_dir = 'data'
    all_items = []
    
    if not os.path.exists(data_dir):
        print("错误: 找不到data目录")
        return
    
    # 获取所有JSON文件
    json_files = [f for f in os.listdir(data_dir) if f.endswith('.json')]
    if not json_files:
        print("错误: data目录中没有JSON文件")
        return
    
    if use_latest_only:
        # 只使用最新的文件
        json_files = [max(json_files, key=lambda x: os.path.getmtime(os.path.join(data_dir, x)))]
        print(f"只处理最新文件: {json_files[0]}")
    
    for filename in json_files:
        file_path = os.path.join(data_dir, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                items = data.get('items', [])
                all_items.extend(items)
                print(f"从 {filename} 读取了 {len(items)} 条电影数据")
        except Exception as e:
            print(f"警告: 读取文件 {filename} 时出错: {e}")
            continue
    
    if not all_items:
        print("错误: 没有找到电影数据")
        return
    
    # 准备数据列表
    movies_data = []
    
    for item in all_items:
        # 解析副标题信息
        subtitle = item.get('card_subtitle', '')
        parts = subtitle.split(' / ')
        
        # 提取各部分信息
        country = parts[1] if len(parts) > 1 else ''
        movie_type = parts[2] if len(parts) > 2 else ''
        director = parts[3] if len(parts) > 3 else ''
        actors = parts[4] if len(parts) > 4 else ''
        
        # 提取关键信息
        movie_info = {
            '电影ID': item.get('id', ''),
            '电影标题': item.get('title', ''),
            '年份': item.get('year', ''),
            '评分': item.get('rating', {}).get('value', 0),
            '评分人数': item.get('rating', {}).get('count', 0),
            '制片国家': country,
            '影片类型': movie_type,
            '导演': director,
            '主演': actors,
            '封面链接': item.get('pic', {}).get('normal', '') if item.get('pic') else ''
        }
        movies_data.append(movie_info)
    
    # 创建DataFrame
    df = pd.DataFrame(movies_data)
    
    # 读取配置文件获取tags参数
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        tags = config.get('tags', '')
    except Exception:
        tags = ''
    
    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"豆瓣电影{tags}年电影合集报表_{timestamp}.xlsx"
    
    # 创建exports目录
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    export_path = os.path.join(export_dir, excel_filename)
    
    # 使用openpyxl创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "豆瓣电影"
    
    # 写入表头
    headers = list(df.columns) + ['封面']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # 设置表头样式
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据
    for row_num, (_, row) in enumerate(df.iterrows(), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # 设置数据样式
            if col_num == 4:  # 评分列
                if isinstance(value, (int, float)) and value >= 8.0:
                    cell.font = Font(bold=True, color="E74C3C")
                elif isinstance(value, (int, float)) and value >= 7.0:
                    cell.font = Font(bold=True, color="F39C12")
            
            # 列启用自动换行
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 下载并插入图片（如果启用）
    if include_images:# 下载并插入图片（如果启用）
        print("正在下载封面图片...")
        image_count = 0
        failed_count = 0
        max_image_width = 0  # 记录最大图片宽度
        
        # 计算需要下载的图片总数
        total_images = sum(1 for _, row in df.iterrows() if row['封面链接'])
        print(f"总共需要下载 {total_images} 张图片")
        
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            image_url = row['封面链接']
            if image_url:
                # 显示下载进度 - 始终显示进度，包括第一张图片
                progress = (image_count / total_images) * 100 if total_images > 0 else 0
                print(f"正在下载第 {image_count + 1}/{total_images} 张图片 ({progress:.1f}%)")
                
                # 添加额外的进度提示，避免长时间无输出
                if image_count % 5 == 0 and image_count > 0:
                    print(f"进度: 已处理 {image_count}/{total_images} 张图片 ({progress:.1f}%)")
                
                img_data = download_image(image_url)
                if img_data:
                    try:
                        img = Image(img_data)
                        # 计算等比例缩放后的尺寸（固定宽度90像素，高度按比例缩放）
                        scale_factor = 90 / img.width
                        target_height = img.height * scale_factor
                        # 设置图片宽度为90像素，高度按比例缩放
                        img.width = 90
                        img.height = target_height
                        
                        # 插入到封面列（最后一列）
                        col_letter = get_column_letter(len(headers))
                        cell_ref = f"{col_letter}{row_num}"
                        ws.add_image(img, cell_ref)
                        image_count += 1
                        
                        # 更新最大图片宽度
                        max_image_width = max(max_image_width, img.width)
                        
                        # 根据用户要求，有封面的行高设置为固定96磅
                        ws.row_dimensions[row_num].height = 96
                        
                        # 显示下载成功信息
                        print(f"下载图片 {image_count}/{total_images} - {image_url}")
                        
                    except Exception as e:
                        print(f"插入图片失败 {image_url}: {e}")
                else:
                    print(f"下载失败 {image_count + 1}/{total_images} - {image_url}")
                    failed_count += 1
    else:
        print("跳过封面图片下载")
    
    # 优化列宽自动调整
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        # 考虑中文字符宽度（中文字符占2个英文字符宽度）
                        cell_length = 0
                        for char in str(cell.value):
                            if '一' <= char <= '鿿':  # 中文字符
                                cell_length += 2
                            else:
                                cell_length += 1
                        max_length = max(max_length, cell_length)
                except:
                    pass
        # 设置更合理的列宽，最小10，最大40
        adjusted_width = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # 根据实际图片宽度动态设置封面列宽（加5像素边距确保完全包含）
    cover_col_letter = get_column_letter(len(headers))
    if include_images and max_image_width > 0:
        # 将像素宽度转换为Excel字符宽度（Excel中1字符 ≈ 7像素）
        excel_width = (max_image_width + 5) / 7
        ws.column_dimensions[cover_col_letter].width = excel_width
    else:
        # 如果没有图片，设置默认宽度（95像素 ≈ 13.57字符）
        ws.column_dimensions[cover_col_letter].width = 13.57
    
    # 优化表格样式
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # 设置交替行颜色以提高可读性
    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:  # 偶数行
            for cell in ws[row_num]:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # 冻结窗格（固定表头）
    ws.freeze_panes = "A2"
    
    # 优化表头样式
    ws.row_dimensions[1].height = 30  # 增加表头行高
    
    # 设置表头字体更大更醒目
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=14)
    
    # 优化数据行高设置
    for row_num in range(2, ws.max_row + 1):
        # 只在没有设置过行高的行设置默认行高（没有封面图片的行）
        if ws.row_dimensions[row_num].height is None:
            ws.row_dimensions[row_num].height = 20  # 没有图片的行使用固定行高
            
            # 为包含长文本的单元格启用自动换行
            for cell in ws[row_num]:
                if cell.column_letter in ['G', 'H']:  # 副标题和用户评论列
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # 保存Excel文件
    try:
        wb.save(export_path)
        print(f"导出完成 {len(movies_data)} 条数据")
        if include_images:
            print(f"图片插入 {image_count} 张，失败 {failed_count} 张")
        print("\n数据列包含:")
        for col in headers:
            print(f"- {col}")
    except Exception as e:
        print(f"导出Excel时出错: {e}")

if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='导出豆瓣电影数据到Excel')
    parser.add_argument('--no-images', action='store_true', 
                       help='不下载封面图片，仅保留封面链接')
    parser.add_argument('--all-files', action='store_true',
                       help='处理所有JSON文件，而不仅是最新的')
    parser.add_argument('--progress', action='store_true',
                       help='显示进度信息')
    
    args = parser.parse_args()
    
    export_douban_to_excel(
        use_latest_only=not args.all_files,
        include_images=not args.no_images
    )